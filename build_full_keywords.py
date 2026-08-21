#!/usr/bin/env python3
"""
Build comprehensive competitor keywords Excel with semantic expansion
Sources: Keys.so query reports + web scrape verification + semantic expansion
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import itertools

# === KEYS.SO EXTRACTED DATA ===

# Query: "нейросеть онлайн" - дополняющие фразы + organic top 10
neuro_online_related = [
    ("нейросеть онлайн", 409554, "нейросеть онлайн"),
    ("онлайн нейросеть", 406327, "нейросеть онлайн"),
    ("бесплатная нейросеть онлайн", 301209, "нейросеть онлайн"),
    ("ии онлайн бесплатно", 231309, "нейросеть онлайн"),
    ("онлайн ии", 382060, "нейросеть онлайн"),
    ("нейросеть онлайн бесплатно на русском", 39676, "нейросеть онлайн"),
    ("нейросеть сайт онлайн", 792, "нейросеть онлайн"),
    ("нейронка онлайн без регистрации", 49, "нейросеть онлайн"),
    ("бесплатный ии", 0, "нейросеть онлайн"),
    ("ии нейросеть", 0, "нейросеть онлайн"),
]

# Organic winners for "нейросеть онлайн"
neuro_online_organic = [
    ("giga чат нейросеть", 0, "giga.chat"),
    ("гигачат онлайн", 0, "giga.chat"),
    ("алиса нейросеть", 0, "giga.chat"),
    ("rugpt нейросеть", 0, "rugpt.io"),
    ("robotext нейросеть", 0, "robotext.io"),
    ("textplus нейросеть", 0, "textplus.ru"),
    ("retext нейросеть", 0, "retext.ai"),
    ("geekbot нейросеть", 0, "geekbot.ru"),
    ("sinonim нейросеть", 0, "sinonim.org"),
]

# Query: "чат gpt" (2.38M freq)
chatgpt_related = [
    ("чат gpt", 2382010, "чат gpt"),
    ("чат гпт", 276, "чат gpt"),
    ("чат gpt т", 116, "чат gpt"),
    ("чатгтп", 341, "чат gpt"),
    ("gpt чат", 0, "чат gpt"),
    ("chat gpt на русском", 0, "чат gpt"),
    ("чат gpt бесплатно", 0, "чат gpt"),
    ("chatgpt онлайн", 0, "чат gpt"),
    ("chatgpt без впн", 0, "чат gpt"),
    ("chatgpt россия", 0, "чат gpt"),
    ("gpt чат бот", 0, "чат gpt"),
    ("chatgpt telegram", 0, "чат gpt"),
    ("chatgpt без ограничений", 0, "чат gpt"),
    ("чат с chatgpt", 0, "чат gpt"),
    ("chatgpt диалог", 0, "чат gpt"),
    ("гпт чат", 0, "чат gpt"),
    ("chatgpt 4", 0, "чат gpt"),
    ("chatgpt 4o", 0, "чат gpt"),
    ("gpt 4 чат", 0, "чат gpt"),
]

# Query: "чат gpt" - contextual ads showed
chatgpt_ads = [
    ("гигачат", 0, "giga.chat (ads)"),
    ("gpt в телеграм", 0, "t.me (ads)"),
    ("чат gpt t.me", 0, "t.me (ads)"),
    ("gpt телеграм бот", 0, "t.me (ads)"),
    ("gpt тг бот", 0, "t.me (ads)"),
    ("нейросеть телеграм", 0, "t.me (ads)"),
]

# Query: "deepseek" (197K freq) + "дипсик" (3.14M freq)
deepseek_related = [
    ("deepseek", 197148, "deepseek"),
    ("deepseek чат", 3328, "deepseek"),
    ("deep seek chat", 1981, "deepseek"),
    ("deepseek на русском", 0, "deepseek"),
    ("deepseek без впн", 0, "deepseek"),
    ("deepseek войти", 0, "deepseek"),
    ("deepseek нейросеть", 0, "deepseek"),
    ("deepseek desktop", 0, "deepseek"),
    ("deep seek скачать", 20, "deepseek"),
    ("deepseeker", 0, "deepseek"),
    ("дипсик", 3139258, "дипсик"),
    ("дипсик что это", 2463, "дипсик"),
    ("дипсик ии", 7451, "дипсик"),
    ("дипсик нейросеть", 0, "дипсик"),
    ("дипсик онлайн", 0, "дипсик"),
    ("дипсик чат", 0, "дипсик"),
    ("искусственный интеллект дипсик", 0, "дипсик"),
    ("депсик нейросеть", 0, "дипсик"),
    ("дип сик", 0, "дипсик"),
    ("deep seek", 0, "deepseek"),
    ("dipseek", 16727, "deepseek"),
    ("deek seek", 207, "deepseek"),
    ("http deepseek com", 261, "deepseek"),
]

# DeepSeek related - contextual ads
deepseek_ads = [
    ("нейросеть max.ru", 0, "max.ru (ads)"),
    ("макс нейросеть", 0, "max.ru (ads)"),
    ("max нейросеть deepseek", 0, "max.ru (ads)"),
    ("алиса deepseek", 0, "alice.yandex.ru (ads)"),
    ("агрегатор нейросетей макс", 0, "max.ru (ads)"),
    ("бесплатный доступ нейросетям max", 0, "max.ru (ads)"),
    ("получите доступ нейросетям", 0, "max.ru (ads)"),
    ("редиграм нейросеть", 0, "redigram.ru (ads)"),
    ("redigram ии", 0, "redigram.ru (ads)"),
    ("персональный ии помощник", 0, "redigram.ru (ads)"),
    ("помощник ии redigram", 0, "redigram.ru (ads)"),
]

# Query: "нейросеть чат" (283K freq)
neuro_chat_related = [
    ("нейросеть чат", 283353, "нейросеть чат"),
    ("чат нейросеть", 354757, "нейросеть чат"),
    ("ии нейросеть чат", 1540, "нейросеть чат"),
    ("нейросеть для общения бесплатно", 0, "нейросеть чат"),
    ("чат с нейросетью", 0, "нейросеть чат"),
    ("чат нейросеть онлайн", 0, "нейросеть чат"),
    ("чат ии", 402555, "чат ии"),
    ("ии общение онлайн", 734, "чат ии"),
    ("ии чаты", 415457, "чат ии"),
    ("чат ии онлайн", 0, "чат ии"),
    ("чат с ии", 0, "чат ии"),
    ("чат gpt нейросеть", 0, "чат ии"),
    ("чат бот ии", 0, "чат ии"),
    ("чат бот нейросеть", 0, "нейросеть чат"),
    ("чат ai", 0, "чат ии"),
    ("айти чат", 0, "чат ии"),
]

# Query: нейросеть чат - ads showed
neuro_chat_ads = [
    ("chadgpt", 0, "chadgpt.ru (ads)"),
    ("чад гпт", 0, "chadgpt.ru (ads)"),
    ("chadgpt нейросеть", 0, "chadgpt.ru (ads)"),
    ("чат gpt chadgpt", 0, "chadgpt.ru (ads)"),
    ("chad gpt", 0, "chadgpt.ru (ads)"),
]

# Aggregator / direct competitor keywords
aggregator_keywords = [
    ("агрегатор нейросетей", 0, "общее"),
    ("все нейросети в одном месте", 0, "общее"),
    ("чат нейросетей", 0, "общее"),
    ("gpt агрегатор", 0, "общее"),
    ("нейросеть агрегатор", 0, "общее"),
    ("агрегатор ии", 0, "общее"),
    ("нейросети агрегатор", 0, "общее"),
    ("доступ ко всем нейросетям", 0, "общее"),
    ("нейросети под одной крышей", 0, "общее"),
    ("объединение нейросетей", 0, "общее"),
    ("все модели ии", 0, "общее"),
    ("маркетплейс нейросетей", 0, "общее"),
    ("витрина нейросетей", 0, "общее"),
    ("каталог нейросетей", 0, "общее"),
    ("нейросети без подписок", 0, "общее"),
    ("оплата нейросетей по факту", 0, "общее"),
    ("нейросети с оплатой в рублях", 0, "общее"),
    ("нейросети картой рф", 0, "общее"),
    ("нейросети сбп", 0, "общее"),
    ("нейросети без впн россия", 0, "общее"),
    ("нейросети для россии", 0, "общее"),
    ("чат нейросетей россия", 0, "общее"),
    ("нейросети на русском", 0, "общее"),
    ("русскоязычные нейросети", 0, "общее"),
]

# AI assistant and chat keywords
ai_assistant_kw = [
    ("искусственный интеллект онлайн", 0, "общее"),
    ("ии ассистент", 0, "общее"),
    ("виртуальный помощник ии", 0, "общее"),
    ("ai помощник", 0, "общее"),
    ("ии бот", 0, "общее"),
    ("искусственный интеллект чат", 0, "общее"),
    ("ai чат бот", 0, "общее"),
    ("чат бот ии", 0, "общее"),
    ("ии консультант", 0, "общее"),
    ("нейросеть ассистент", 0, "общее"),
]

# Model-specific keywords
model_keywords = [
    ("chatgpt 4", 0, "модели"),
    ("gpt 4o", 0, "модели"),
    ("gpt 5", 0, "модели"),
    ("claude нейросеть", 0, "модели"),
    ("claude чат", 0, "модели"),
    ("claude sonnet", 0, "модели"),
    ("gemini нейросеть", 0, "модели"),
    ("gemini чат", 0, "модели"),
    ("gemini google", 0, "модели"),
    ("llama нейросеть", 0, "модели"),
    ("mistral нейросеть", 0, "модели"),
    ("dall e нейросеть", 0, "модели"),
    ("midjourney нейросеть", 0, "модели"),
    ("стабильная диффузия", 0, "модели"),
    ("кандинский нейросеть", 0, "модели"),
    ("sora нейросеть", 0, "модели"),
    ("нейросеть для генерации изображений", 0, "модели"),
    ("нейросеть для текста", 0, "модели"),
    ("нейросеть для кода", 0, "модели"),
    ("нейросеть программирование", 0, "модели"),
    ("deepseek v3", 0, "модели"),
    ("deepseek r1", 0, "модели"),
    ("gpt для россии", 0, "модели"),
    ("chatgpt россия карта", 0, "модели"),
]

# Telegram bot keywords
telegram_kw = [
    ("telegram бот нейросеть", 0, "telegram"),
    ("нейросеть в телеграм", 0, "telegram"),
    ("чат гпт телеграм", 0, "telegram"),
    ("бот нейросеть тг", 0, "telegram"),
    ("gpt бот телеграм", 0, "telegram"),
    ("telegram chatgpt", 0, "telegram"),
    ("deepseek телеграм", 0, "telegram"),
    ("бот с chatgpt", 0, "telegram"),
    ("нейросеть тг бот", 0, "telegram"),
    ("ии бот телеграм", 0, "telegram"),
    ("ai telegram bot", 0, "telegram"),
    ("gpt телеграм", 0, "telegram"),
]

# Competitor-specific keywords from domain structure
competitor_pages = {
    "giga.chat": [
        "гигачат", "gigachat", "giga chat", "сбер нейросеть", "кандинский",
        "гигачат войти", "гигачат онлайн", "gigachat вход", "giga chat регистрация",
        "giga chat нейросеть", "сбер гигачат", "гигачат сбер", "gigachat нейросеть",
        "giga chat отзывы", "гигачат отзывы", "сбер искусственный интеллект",
        "giga chat api", "gigachat api", "гигачат для бизнеса",
    ],
    "rugpt.io": [
        "rugpt", "ru gpt", "rugpt нейросеть", "русский gpt", 
        "нейросеть rugpt", "rugpt io", "gpt на русском языке",
        "rugpt вход", "rugpt регистрация", "rugpt отзывы",
        "нейросеть для текста rugpt", "rugpt тексты", "rugpt генерация текста",
        "rugpt бесплатно", "rugpt онлайн", "проверка орфографии rugpt",
    ],
    "robotext.io": [
        "robotext", "роботекст", "robotext нейросеть", "robotext io",
        "robotext вход", "robotext регистрация", "robotext отзывы",
        "robotext рерайт", "robotext текст", "robotext редактор",
        "роботекст нейросеть", "robotext бесплатно",
    ],
    "geekbot.ru": [
        "geekbot", "гикбот", "geekbot нейросеть", "geek bot",
        "geekbot вход", "geekbot регистрация", "geekbot отзывы",
        "geekbot исправить текст", "fix text ai geekbot",
        "geekbot бесплатно", "geekbot тексты",
    ],
    "chadgpt.ru": [
        "chadgpt", "чад gpt", "chadgpt нейросеть", "chad gpt",
        "chadgpt вход", "chadgpt регистрация", "chadgpt отзывы",
        "чат gpt chadgpt", "chadgpt бесплатно", "chadgpt на русском",
        "chadgpt chat", "chadgpt без впн",
    ],
    "max.ru": [
        "max нейросеть", "max ru нейросеть", "max.deepseek",
        "макс нейросеть", "max нейросети", "агрегатор нейросетей max",
        "max нейросеть вход", "max нейросеть отзывы", "макс нейросеть отзывы",
    ],
    "redigram.ru": [
        "redigram", "редиграм", "redigram нейросеть", "ии помощник redigram",
        "redigram вход", "redigram отзывы", "redigram ии",
        "redigram помощник", "redigram бесплатно",
    ],
    "cropmedia.ru": [
        "cropmedia", "кропмедиа", "cropmedia нейросеть", "cropmedia вход",
        "cropmedia отзывы", "cropmedia агрегатор",
    ],
    "textplus.ru": [
        "textplus", "текст плюс", "textplus нейросеть", "textplus рерайт",
        "textplus вход", "textplus отзывы", "рерайт текста textplus",
    ],
    "retext.ai": [
        "retext", "retext ai", "retext нейросеть", "retext перефразирование",
        "retext вход", "retext отзывы", "retext бесплатно",
        "retext ai текст", "retext рерайт",
    ],
    "sinonim.org": [
        "sinonim", "синоним", "синонимайзер", "синонимизатор",
        "sinonim org", "синоним нейросеть",
    ],
}

# === COMBINE ALL ===
all_keywords = []

def add_kw(kw_list, source_label):
    for item in kw_list:
        kw, freq, src = item if len(item) == 3 else (item[0], item[1], source_label)
        all_keywords.append({
            'keyword': kw,
            'frequency': freq,
            'source': src
        })

# Add from all sources
add_kw(neuro_online_related, "нейросеть онлайн (related)")
add_kw(neuro_online_organic, "нейросеть онлайн (organic)")
add_kw(chatgpt_related, "чат gpt")
add_kw(chatgpt_ads, "чат gpt (ads)")
add_kw(deepseek_related, "deepseek/дипсик")
add_kw(deepseek_ads, "deepseek (ads)")
add_kw(neuro_chat_related, "нейросеть чат/чат ии")
add_kw(neuro_chat_ads, "нейросеть чат (ads)")
add_kw(aggregator_keywords, "агрегаторы")
add_kw(ai_assistant_kw, "ии ассистенты")
add_kw(model_keywords, "модели")
add_kw(telegram_kw, "telegram")

# Add competitor-specific keywords
for site, kws in competitor_pages.items():
    for kw in kws:
        all_keywords.append({
            'keyword': kw,
            'frequency': 0,
            'source': f'competitor: {site}'
        })

# === CLUSTERING ===
clusters = {
    "Агрегаторы нейросетей (прямые конкуренты)": [
        "агрегатор", "все нейросети в одном", "чат нейросетей", "нейросеть агрегатор",
        "gpt агрегатор", "агрегатор ии", "все модели ии", "маркетплейс нейросетей",
        "витрина нейросетей", "каталог нейросетей", "объединение нейросетей",
        "нейросети под одной крышей", "доступ ко всем нейросетям",
        "giga.chat", "gigachat", "гигачат", "сбер нейросеть", "giga chat",
        "chadgpt", "чад gpt", "chad gpt", "chadgpt без впн", "chadgpt на русском",
        "geekbot", "гикбот", "geek bot", "geekbot нейросеть",
        "max нейросеть", "макс нейросеть", "max.deepseek", "агрегатор нейросетей max",
        "max нейросети", "cropmedia", "кропмедиа", "redigram", "редиграм",
    ],
    "ChatGPT и GPT": [
        "чат gpt", "чат гпт", "gpt чат", "гпт чат", "чат gpt т", "чатгтп",
        "chat gpt на русском", "чат gpt бесплатно", "chatgpt онлайн",
        "chatgpt без впн", "chatgpt россия", "chatgpt telegram",
        "chatgpt без ограничений", "чат с chatgpt", "chatgpt диалог",
        "chatgpt 4", "chatgpt 4o", "gpt 4 чат", "gpt 4o",
        "gpt 5", "rugpt", "ru gpt", "русский gpt", "gpt на русском языке",
        "rugpt нейросеть", "rugpt io", "rugpt тексты", "chadgpt нейросеть",
        "gpt для россии", "chatgpt россия карта",
    ],
    "DeepSeek / Дипсик": [
        "deepseek", "deepseek чат", "deep seek chat", "deepseek на русском",
        "deepseek без впн", "deepseek войти", "deepseek нейросеть",
        "deepseek desktop", "deep seek скачать", "deepseeker",
        "dipseek", "deek seek", "http deepseek com",
        "дипсик", "дипсик что это", "дипсик ии", "дипсик нейросеть",
        "дипсик онлайн", "дипсик чат", "депсик нейросеть", "дип сик",
        "искусственный интеллект дипсик", "deepseek v3", "deepseek r1",
        "deepseek телеграм",
    ],
    "Нейросеть онлайн (общие запросы)": [
        "нейросеть онлайн", "онлайн нейросеть", "бесплатная нейросеть онлайн",
        "нейросеть онлайн бесплатно на русском", "нейросеть сайт онлайн",
        "нейронка онлайн без регистрации",
        "ии онлайн бесплатно", "онлайн ии", "бесплатный ии", "ии нейросеть",
    ],
    "Чат с ИИ / нейросеть чат": [
        "нейросеть чат", "чат нейросеть", "ии нейросеть чат",
        "нейросеть для общения бесплатно", "чат с нейросетью",
        "чат нейросеть онлайн", "чат бот нейросеть",
        "чат ии", "ии общение онлайн", "ии чаты",
        "чат ии онлайн", "чат с ии", "чат gpt нейросеть",
        "чат бот ии", "чат ai", "айти чат",
        "искусственный интеллект чат", "ai чат бот",
    ],
    "Нейросети для России (без VPN / рублёвая оплата)": [
        "нейросети без впн россия", "нейросети для россии",
        "чат нейросетей россия", "нейросети на русском",
        "русскоязычные нейросети", "нейросети без впн",
        "нейросети с оплатой в рублях", "нейросети картой рф",
        "нейросети сбп", "оплата нейросетей по факту",
        "нейросети без подписок", "chatgpt без впн",
        "deepseek на русском", "deepseek без впн",
        "chatgpt россия", "gpt для россии",
    ],
    "Telegram-боты с нейросетями": [
        "telegram бот нейросеть", "нейросеть в телеграм",
        "чат гпт телеграм", "бот нейросеть тг",
        "gpt бот телеграм", "telegram chatgpt",
        "бот с chatgpt", "нейросеть тг бот",
        "ии бот телеграм", "ai telegram bot",
        "gpt телеграм", "gpt в телеграм",
        "gpt телеграм бот", "gpt тг бот",
        "нейросеть телеграм", "deepseek телеграм",
    ],
    "ИИ-помощники и ассистенты": [
        "искусственный интеллект онлайн", "ии ассистент",
        "виртуальный помощник ии", "ai помощник",
        "ии бот", "ии консультант",
        "нейросеть ассистент", "персональный ии помощник",
        "помощник ии redigram", "redigram ии",
        "алиса нейросеть", "алиса", "алиса нейросеть онлайн",
    ],
    "Модели и инструменты (Claude, Gemini, Llama...)": [
        "claude нейросеть", "claude чат", "claude sonnet",
        "gemini нейросеть", "gemini чат", "gemini google",
        "llama нейросеть", "mistral нейросеть",
        "dall e нейросеть", "midjourney нейросеть",
        "стабильная диффузия", "кандинский нейросеть",
        "sora нейросеть", "нейросеть для генерации изображений",
        "нейросеть для кода", "нейросеть программирование",
    ],
    "Текстовые нейросети / рерайт / редакторы": [
        "robotext", "роботекст", "robotext нейросеть", "robotext io",
        "robotext рерайт", "robotext текст", "robotext редактор",
        "текст плюс", "textplus", "textplus нейросеть", "textplus рерайт",
        "retext", "retext ai", "retext нейросеть", "retext перефразирование",
        "retext ai текст", "retext рерайт",
        "sinonim", "синоним", "синонимайзер", "синонимизатор",
        "нейросеть для текста", "нейросеть тексты", "геекбот исправить текст",
        "fix text ai", "geekbot исправить текст", "проверка орфографии",
    ],
}

# === DEDUPLICATE AND CLUSTER ===
seen = set()
clustered = defaultdict(list)

for item in all_keywords:
    kw_lower = item['keyword'].lower().strip()
    if kw_lower in seen:
        continue
    seen.add(kw_lower)
    
    # Find cluster
    assigned = "Прочее"
    for cl_name, cl_kws in clusters.items():
        for cl_kw in cl_kws:
            if cl_kw.lower() in kw_lower or kw_lower in cl_kw.lower():
                assigned = cl_name
                break
        if assigned != "Прочее":
            break
    
    clustered[assigned].append(item)
    item['cluster'] = assigned

# Sort by frequency
all_items = sorted(all_keywords, key=lambda x: x['frequency'], reverse=True)

# Also deduplicate from seen
deduped = []
seen2 = set()
for item in all_items:
    if item['keyword'].lower().strip() not in seen2:
        seen2.add(item['keyword'].lower().strip())
        deduped.append(item)

# === BUILD EXCEL ===
wb = Workbook()

# -- Sheet 1: All keywords --
ws = wb.active
ws.title = "Все ключевые слова"
headers = ['Ключевое слово', 'Частотность', 'Кластер', 'Источник']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

cluster_fills = {}
cluster_colors = ['D6E4F0', 'E2EFDA', 'FCE4D6', 'D9E2F3', 'EDEDED', 'FFF2CC', 'E2D9F3', 'D6F0E4', 'F8D7DA', 'D1ECF1', 'FCE4EC']
for i, (cl_name, _) in enumerate(sorted(clustered.items())):
    cluster_fills[cl_name] = PatternFill(start_color=cluster_colors[i % len(cluster_colors)], 
                                          end_color=cluster_colors[i % len(cluster_colors)], fill_type='solid')

cluster_fonts = {}
for cl_name in clustered:
    cluster_fonts[cl_name] = Font(bold=True, color='2F5496')

for row_idx, item in enumerate(deduped, 2):
    ws.cell(row=row_idx, column=1, value=item['keyword']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=row_idx, column=2, value=item['frequency']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell_c = ws.cell(row=row_idx, column=3, value=item['cluster'])
    cell_c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    if item['cluster'] in cluster_fills:
        cell_c.fill = cluster_fills[item['cluster']]
    ws.cell(row=row_idx, column=4, value=item['source']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 30
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:D{len(deduped)+1}"

# -- Sheet 2: By cluster --
ws2 = wb.create_sheet("По кластерам")
for col, h in enumerate(['Кластер', 'Количество'], 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

row = 2
for cl_name in sorted(clustered.keys()):
    kws_in_cluster = [x['keyword'] for x in deduped if x['cluster'] == cl_name]
    ws2.cell(row=row, column=1, value=cl_name).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws2.cell(row=row, column=1).fill = cluster_fills.get(cl_name, PatternFill())
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=len(kws_in_cluster)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    row += 1

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15

# -- Sheet 3: By competitor --
ws3 = wb.create_sheet("По конкурентам")
for col, h in enumerate(['Конкурент', 'Ключевые слова'], 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

row = 2
for site, kws in sorted(competitor_pages.items()):
    ws3.cell(row=row, column=1, value=site).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws3.cell(row=row, column=1).fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=2, value='; '.join(kws)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws3.cell(row=row, column=2).value = '; '.join(kws)
    row += 1

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 100

# -- Sheet 4: Comparison --
ws4 = wb.create_sheet("Сравнение")
headers4 = ['Параметр', 'giga.chat', 'rugpt.io', 'robotext.io', 'geekbot.ru', 'chadgpt.ru', 'ai-sphere.ru']
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

comp_data = [
    ('Трафик органика', '~21 958', '~4 295', '~7 478', '~1 975', '~393', '0'),
    ('Запросов топ-50', '43 671', '64 570', '32 273', '21 877', '5 840', '0'),
    ('Страниц в индексе', '355', '422', '91', '612', '23', '0'),
    ('DR', '39', '28', '35', '25', '21', '0'),
    ('Бюджет Директа', '~37 млн ₽', '~9,9 млн ₽', '~4,3 млн ₽', '—', '~7 млн ₽', '~81 ₽'),
    ('Объявлений', '524', '1 166', '206', '—', '220', '2'),
    ('Запросов в контексте', '13 732', '8 044', '2 476', '—', '3 327', '2'),
    ('ИИ-упоминаний Алисы', '5 556', '619', '769', '535', '52', '0'),
]
for row_idx, (param, *vals) in enumerate(comp_data, 2):
    ws4.cell(row=row_idx, column=1, value=param).font = Font(bold=True)
    ws4.cell(row=row_idx, column=1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, v in enumerate(vals, 2):
        c = ws4.cell(row=row_idx, column=ci, value=v)
        c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        c.alignment = Alignment(horizontal='center')

ws4.column_dimensions['A'].width = 30
for ci in range(2, 8):
    ws4.column_dimensions[get_column_letter(ci)].width = 18

# Save
path = '/root/ai-sphere/competitors_keywords_full.xlsx'
wb.save(path)

total_unique = len(deduped)
print(f"Excel: {path}")
print(f"Всего уникальных ключевых слов: {total_unique}")
print(f"Кластеров: {len([c for c in clustered if clustered[c]])}")
for cl_name in sorted(clustered.keys()):
    cnt = len([x for x in deduped if x['cluster'] == cl_name])
    if cnt > 0:
        print(f"  {cl_name}: {cnt}")
