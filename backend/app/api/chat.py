"""Chat API: model listing, chat completion proxy via OpenRouter, file upload."""

import json
import re
import logging
import math
import uuid
import io
import time
import unicodedata
import asyncio
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from sqlalchemy import select, update, case, delete

import httpx
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request, status
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user, get_optional_user
from app.core.product_events import record_server_event
from app.core.economics import credits_for_provider_cost, pricing_context, enforce_free_program_budget
from app.core.limiter import limiter
from app.models.user import User
from app.models.transaction import Transaction
from app.models.ai_model import AiModel
from app.models.user_query import UserQuery
from app.models.feedback import MessageFeedback
from app.models.file_record import FileRecord
from app.models.chat_session import ChatSession as ChatSessionModel
from app.models.credit_op import CreditOperation
from app.core.credits import allocate_buckets, bucket_snapshot
from app.schemas.chat import (
    ChatRequest, ModelInfo, SessionSaveRequest, SessionResponse, FactCheckRequest,
    FactCheckResponse, FactCheckClaim, MessageFeedbackRequest,
    VoicePunctuateRequest, VoicePunctuateResponse,
)
from app.api.web_search import web_search, needs_search

logger = logging.getLogger("ai-sphere.chat")

router = APIRouter(prefix="/api/chat", tags=["chat"])

VOICE_PUNCTUATION_PROMPT = """Расставь знаки препинания и заглавные буквы в расшифровке речи.
Нельзя добавлять, удалять, заменять или переставлять слова и числа.
Не исправляй грамматику, оговорки и слова-паразиты.
Верни только обработанный текст без пояснений и Markdown."""


def _spoken_tokens(value: str) -> list[str]:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return re.findall(r"[^\W_]+", normalized, flags=re.UNICODE)


def _safe_punctuation_result(source: str, candidate: str) -> bool:
    """Accept punctuation/case changes only; every spoken token must remain intact."""
    return bool(candidate.strip()) and _spoken_tokens(source) == _spoken_tokens(candidate)


def _supports_text_output(model: AiModel) -> bool:
    try:
        return "text" in set(json.loads(model.output_modalities or "[]"))
    except (TypeError, json.JSONDecodeError):
        return False


async def _voice_punctuation_model(db: AsyncSession) -> AiModel | None:
    preferred = (await db.execute(select(AiModel).where(
        AiModel.or_model_id == settings.voice_punctuation_model,
        AiModel.is_active == True,
        AiModel.is_visible == True,
    ))).scalar_one_or_none()
    if preferred and _supports_text_output(preferred):
        return preferred

    candidates = (await db.execute(select(AiModel).where(
        AiModel.is_active == True,
        AiModel.is_visible == True,
    ))).scalars().all()
    text_models = [model for model in candidates if _supports_text_output(model)]
    if not text_models:
        return None
    return min(text_models, key=lambda model: (
        max(0.0, model.or_input_cost or 0.0) + max(0.0, model.or_output_cost or 0.0),
        model.sort_order,
        model.name,
    ))


def _voice_fallback(source: str, started_at: float, model: str, reason: str) -> VoicePunctuateResponse:
    logger.info(
        "Voice punctuation status=fallback model=%s chars=%d duration_ms=%d reason=%s",
        model or "none", len(source), round((time.monotonic() - started_at) * 1000), reason,
    )
    return VoicePunctuateResponse(result=source, applied=False)

# ──────────────── File upload ────────────────

UPLOAD_DIR = settings.uploads_dir
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB
ALLOWED_TYPES = {
    "image/png", "image/jpeg", "image/webp", "image/gif",
    "video/mp4", "video/mpeg", "video/quicktime", "video/webm",
    "application/pdf",
    "text/plain", "text/csv",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
ALLOWED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".mp4", ".mpeg", ".mov", ".webm", ".pdf", ".txt", ".csv", ".docx", ".xlsx"}


def _signature_matches(content: bytes, ext: str) -> bool:
    signatures = {
        ".png": (b"\x89PNG\r\n\x1a\n",), ".jpg": (b"\xff\xd8\xff",), ".jpeg": (b"\xff\xd8\xff",),
        ".gif": (b"GIF87a", b"GIF89a"), ".webp": (b"RIFF",), ".pdf": (b"%PDF-",),
        ".docx": (b"PK\x03\x04",), ".xlsx": (b"PK\x03\x04",),
    }
    if ext in {".txt", ".csv"}:
        return b"\x00" not in content[:4096]
    if ext in {".mp4", ".mov"}:
        return len(content) >= 12 and content[4:8] == b"ftyp"
    if ext in {".mpeg"}:
        return content.startswith((b"\x00\x00\x01", b"ID3"))
    if ext == ".webm":
        return content.startswith(b"\x1aE\xdf\xa3")
    expected = signatures.get(ext)
    return bool(expected and content.startswith(expected))


def _extract_document_text(content: bytes, ext: str) -> str:
    """Extract bounded plain text from the document formats advertised by UI."""
    if ext in {".txt", ".csv"}:
        return content.decode("utf-8", errors="replace")[:100_000]
    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        return "\n".join((page.extract_text() or "") for page in reader.pages)[:100_000]
    if ext == ".docx":
        from docx import Document
        document = Document(io.BytesIO(content))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)[:100_000]
    if ext == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[{sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                lines.append("\t".join("" if value is None else str(value) for value in row))
                if sum(len(line) for line in lines) > 100_000:
                    break
        return "\n".join(lines)[:100_000]
    return ""


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    session_id: str | None = Form(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload a file (max 20 MB). Returns file metadata."""
    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="Файл слишком большой. Максимум 20 МБ")

    # Basic extension-based type detection
    ext = Path(file.filename or "file").suffix.lower()
    if ext not in ALLOWED_EXTENSIONS or (file.content_type and file.content_type not in ALLOWED_TYPES):
        raise HTTPException(status_code=415, detail="Неподдерживаемый формат файла")
    if not _signature_matches(content, ext):
        raise HTTPException(status_code=415, detail="Содержимое файла не соответствует его формату")
    if session_id:
        session = await db.get(ChatSessionModel, session_id)
        if session is not None and session.user_id != user.id:
            raise HTTPException(status_code=403, detail="Чат принадлежит другому пользователю")
    file_id = str(uuid.uuid4())
    saved_name = f"{file_id}{ext}"
    save_path = UPLOAD_DIR / saved_name

    try:
        extracted_text = _extract_document_text(content, ext)
    except Exception as exc:
        logger.warning("Document extraction failed for %s: %s", file.filename, exc)
        raise HTTPException(status_code=422, detail="Не удалось безопасно прочитать документ") from exc

    save_path.write_bytes(content)
    record = FileRecord(
        user_id=user.id, chat_id=session_id or None, filename=saved_name,
        original_name=(file.filename or "unnamed")[:255], mime_type=file.content_type,
        size_bytes=len(content), storage_path=str(save_path), status="processed",
        expires_at=datetime.now(timezone.utc) + timedelta(days=30),
    )
    db.add(record)
    try:
        await db.commit()
        await db.refresh(record)
    except Exception:
        save_path.unlink(missing_ok=True)
        raise

    return {
        "id": str(record.id),
        "file_id": str(record.id),
        "name": file.filename or "unnamed",
        "size": len(content),
        "type": file.content_type or "application/octet-stream",
        "url": f"/api/chat/files/{record.id}",
        "extracted_text": extracted_text,
        "session_id": session_id,
        "expires_at": record.expires_at.isoformat(),
    }


@router.get("/files/{file_id}")
async def download_file(
    file_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    record = (await db.execute(select(FileRecord).where(
        FileRecord.id == file_id, FileRecord.user_id == user.id,
        FileRecord.deleted_at.is_(None), FileRecord.is_blocked == False,
    ))).scalar_one_or_none()
    if record is None:
        raise HTTPException(404, "Файл не найден")
    now = datetime.now(timezone.utc)
    expiry = record.expires_at
    if expiry and (expiry if expiry.tzinfo else expiry.replace(tzinfo=timezone.utc)) <= now:
        raise HTTPException(410, "Срок хранения файла истёк")
    target = Path(record.storage_path).resolve()
    root = settings.uploads_dir.resolve()
    if root not in target.parents or not target.is_file():
        raise HTTPException(404, "Файл не найден")
    return FileResponse(target, media_type=record.mime_type, filename=record.original_name)


@router.post("/voice/punctuate", response_model=VoicePunctuateResponse)
@limiter.limit("20/minute")
async def punctuate_voice(
    request: Request,
    req: VoicePunctuateRequest,
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add punctuation to a voice transcript without changing its spoken tokens."""
    started_at = time.monotonic()
    source = req.text
    if not settings.openrouter_api_key:
        return _voice_fallback(source, started_at, "", "missing_api_key")

    model = await _voice_punctuation_model(db)
    if not model:
        return _voice_fallback(source, started_at, "", "no_text_model")

    model_id = model.or_model_id
    headers = {
        "Authorization": f"Bearer {settings.openrouter_api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://ai-sphere.ru",
        "X-Title": "AI-Sphere",
    }
    body = {
        "model": model_id,
        "messages": [
            {"role": "system", "content": VOICE_PUNCTUATION_PROMPT},
            {"role": "user", "content": source},
        ],
        "max_tokens": 1024,
        "temperature": 0,
    }

    try:
        async with asyncio.timeout(4.0):
            async with httpx.AsyncClient(timeout=4.0, proxy=settings.openrouter_proxy or None) as client:
                response = await client.post(
                    f"{settings.openrouter_base_url}/chat/completions",
                    headers=headers,
                    json=body,
                )
        if response.status_code != 200:
            return _voice_fallback(source, started_at, model_id, f"provider_{response.status_code}")
        payload = response.json()
        candidate = str(payload["choices"][0]["message"].get("content") or "").strip()
        if not _safe_punctuation_result(source, candidate):
            return _voice_fallback(source, started_at, model_id, "unsafe_result")
    except Exception as exc:
        return _voice_fallback(source, started_at, model_id, type(exc).__name__)

    applied = candidate != source
    logger.info(
        "Voice punctuation status=success model=%s chars=%d duration_ms=%d applied=%s",
        model_id, len(source), round((time.monotonic() - started_at) * 1000), applied,
    )
    return VoicePunctuateResponse(result=candidate, applied=applied)



def _to_model_info(model: AiModel) -> ModelInfo:
    def parse(value: str, fallback):
        try:
            return json.loads(value)
        except (TypeError, json.JSONDecodeError):
            return fallback

    return ModelInfo(
        id=model.or_model_id,
        name=model.name,
        provider=model.provider,
        price_per_1k_input=max(0.0, float(model.price_input)),
        price_per_1k_output=max(0.0, float(model.price_output)),
        fixed_price=max(0.0, float(model.fixed_price)),
        context_window=model.max_context or model.max_input_tokens,
        vision=bool(model.vision),
        image_generation_only=bool(model.fixed_price and not model.vision),
        is_active=bool(model.is_active),
        is_visible=bool(model.is_visible),
        input_modalities=parse(model.input_modalities, ["text"]),
        output_modalities=parse(model.output_modalities, ["text"]),
        supported_parameters=parse(model.supported_parameters, {}),
        auto_route_enabled=bool(model.auto_route_enabled),
    )


async def _charge_credits(db: AsyncSession, user_id: int, amount: int, description: str) -> bool:
    """Atomically deduct bucketed credits without allowing a negative balance."""
    if amount <= 0:
        return True
    locked = (await db.execute(select(User).where(User.id == user_id).with_for_update())).scalar_one_or_none()
    if locked is None:
        await db.rollback()
        return False
    try:
        allocation = allocate_buckets(bucket_snapshot(locked), amount)
    except ValueError:
        await db.rollback()
        return False
    before = locked.credits
    for bucket, spent in allocation.items():
        if not spent:
            continue
        setattr(locked, f"credits_{bucket}", getattr(locked, f"credits_{bucket}") - spent)
        db.add(CreditOperation(
            user_id=user_id, op_type="spend", credit_type=bucket, amount=-spent,
            balance_before=before, balance_after=before - amount,
            source=description[:255], comment=description,
        ))
    locked.total_spent_credits += amount
    locked.request_count += 1
    db.add(Transaction(user_id=user_id, amount=-amount, type="spend", description=description))
    await db.commit()
    return True


@router.get("/models", response_model=list[ModelInfo])
async def get_models(db: AsyncSession = Depends(get_db)):
    """Return available models with pricing."""
    result = await db.execute(
        select(AiModel)
        .where(AiModel.is_active == True, AiModel.is_visible == True)
        .order_by(AiModel.sort_order, AiModel.name)
    )
    return [_to_model_info(model) for model in result.scalars().all()]


def _prepare_message(msg, input_modalities: list[str]) -> dict:
    """Strip unsupported historic media while preserving accompanying text."""
    content = msg.content
    if isinstance(content, list):
        allowed_types = {"text"}
        if "image" in input_modalities:
            allowed_types.add("image_url")
        if "video" in input_modalities:
            allowed_types.add("video_url")
        content = [part for part in content if isinstance(part, dict) and part.get("type") in allowed_types]
        if not content:
            content = "(вложение недоступно выбранной модели)"
    return {"role": msg.role, "content": content}


@router.post("/completions")
async def chat_completion(
    req: ChatRequest,
    user: User | None = Depends(get_optional_user),
    db: AsyncSession = Depends(get_db),
):
    """Proxy chat completion to OpenRouter. Requires auth."""
    model = (await db.execute(
        select(AiModel).where(AiModel.or_model_id == req.model, AiModel.is_active == True)
    )).scalar_one_or_none()
    if model is None:
        raise HTTPException(status_code=400, detail=f"Unknown model: {req.model}")
    model_info = _to_model_info(model)
    fallback_ids = list(dict.fromkeys(item for item in req.fallback_models if item and item != req.model))[:3]
    fallback_records = (await db.execute(
        select(AiModel).where(
            AiModel.or_model_id.in_(fallback_ids),
            AiModel.is_active == True,
            AiModel.is_visible == True,
        )
    )).scalars().all() if fallback_ids else []
    fallback_map = {item.or_model_id: item for item in fallback_records}
    candidate_models = [model] + [fallback_map[item] for item in fallback_ids if item in fallback_map]

    # Check if model supports vision — only check the last user message
    for m in reversed(req.messages):
        if m.role == "user":
            if isinstance(m.content, list):
                has_images = any(
                    isinstance(part, dict) and part.get("type") == "image_url"
                    for part in m.content
                )
                if has_images and model_info.image_generation_only:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Модель «{model_info.name}» создаёт изображения из текста, но не умеет редактировать загруженные фото. "
                               f"Используйте GPT-4o, Claude, Gemini 2.5 Flash или другую модель с пониманием изображений."
                    )
                if has_images and "image" not in model_info.input_modalities:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Модель «{model_info.name}» не поддерживает изображения. "
                               f"Выберите модель с поддержкой vision: GPT-4o, Claude, Gemini, Llama Vision или другую."
                    )
                has_videos = any(
                    isinstance(part, dict) and part.get("type") == "video_url"
                    for part in m.content
                )
                if has_videos and "video" not in model_info.input_modalities:
                    raise HTTPException(status_code=400, detail=f"Модель «{model_info.name}» не поддерживает анализ видео")
            break  # only check the latest user message

    # Block guests — auth required
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация для отправки сообщений")
    await enforce_free_program_budget(db, user)

    # Credit check for authenticated users
    if user and user.credits <= 0 and (model_info.price_per_1k_input > 0 or model_info.fixed_price > 0):
        raise HTTPException(status_code=402, detail="Недостаточно кредитов. Пополните баланс.")

    # Estimate cost (for list content, count text parts only)
    estimated_input_tokens = 0
    for m in req.messages:
        if isinstance(m.content, str):
            estimated_input_tokens += len(m.content) // 4
        elif isinstance(m.content, list):
            for part in m.content:
                if isinstance(part, dict) and part.get("type") == "text":
                    estimated_input_tokens += len(part.get("text", "")) // 4
    estimated_output_tokens = min(req.max_tokens, model_info.context_window)
    if model_info.fixed_price > 0:
        estimated_cost = model_info.fixed_price
    else:
        estimated_cost = (estimated_input_tokens * model_info.price_per_1k_input / 1000) + (estimated_output_tokens * model_info.price_per_1k_output / 1000)

    if user and user.credits < estimated_cost and (model_info.price_per_1k_input > 0 or model_info.fixed_price > 0):
        raise HTTPException(status_code=402, detail="Недостаточно кредитов для этого запроса")

    # ── Web search (актуальные данные из интернета) ─────────────────
    search_context = ""
    last_query = ""
    try:
        user_msgs = [m for m in req.messages if getattr(m, "role", "") == "user"]
        if user_msgs and needs_search(req.messages):
            last_query = getattr(user_msgs[-1], "content", "")
            if isinstance(last_query, str) and len(last_query) > 10:
                search_context = await web_search(last_query[:200], max_results=5)
                if search_context:
                    logger.info("Web search returned %d chars for: %s", len(search_context), last_query[:60])
                else:
                    logger.info("Web search returned empty for: %s", last_query[:60])
    except Exception as exc:
        logger.warning("Web search error: %s", exc)

    # Фильтр релевантности
    if search_context and isinstance(last_query, str) and len(last_query) > 10:
        topic_words = {w.lower() for w in last_query.split() if len(w) > 3}
        items = search_context.split("\n**")
        relevant_count = 0
        for item in items:
            item_lower = item.lower()
            if any(tw in item_lower for tw in topic_words):
                relevant_count += 1
        if relevant_count < 2:
            logger.info("Search filtered: only %d/%d items relevant", relevant_count, len(items))
            search_context = ""

    headers = {
        "Authorization": f"Bearer {settings.openrouter_api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://ai-sphere.ru",
        "X-Title": "AI-Sphere",
    }

    # System prompt with model identity
    system_prompt = (
        f"Ты — {model_info.name}, ИИ-ассистент от {model_info.provider}. "
        f"Твоё имя — {model_info.name}, тебя создала компания {model_info.provider}. "
        f"Ты НЕ ChatGPT, НЕ GPT, НЕ OpenAI и НЕ ассистент от OpenAI. "
        "Никогда не называй себя ChatGPT, GPT или ассистентом OpenAI. "
        "Отвечай на языке пользователя. Будь полезным, точным и вежливым. "
        f"Сегодня {date.today().strftime('%d.%m.%Y')}. Учитывай актуальную дату в ответах."
    )
    if search_context:
        system_prompt += (
            "\n\nВот актуальная информация из интернета (используй её, если она относится к вопросу пользователя):\n"
            f"{search_context}\n"
            "Если среди результатов поиска нет релевантной информации — отвечай из своих знаний, не выдумывай."
        )

    base_body = {
        "model": req.model,
        "messages": [
            {"role": "system", "content": system_prompt},
        ] + [_prepare_message(m, model_info.input_modalities) for m in req.messages],
        "max_tokens": req.max_tokens,
        "temperature": req.temperature,
    }

    proxy = settings.openrouter_proxy or None

    await record_server_event(
        db, user, "generation_started", template_id=req.template_id,
        task_type=req.task_type or "text", model=req.model,
        metadata={"result_kind": "text"},
    )
    await db.commit()

    # ── Streaming path (skip for image models — images don't stream) ──
    if req.stream and model_info.fixed_price == 0:
        async def event_stream():
            full_content = ""
            final_usage = {}
            effective_model = model
            provider_error = ""
            succeeded = False
            if not settings.openrouter_api_key:
                provider_error = "AI-сервис временно не настроен. Попробуйте позже или обратитесь в поддержку."
            else:
                async with httpx.AsyncClient(timeout=120.0, proxy=proxy) as client:
                    for candidate in candidate_models:
                        body = {
                            **base_body, "model": candidate.or_model_id, "stream": True,
                            "provider": {
                                "sort": "price",
                                "max_price": {
                                    "prompt": max(0.0, candidate.or_input_cost),
                                    "completion": max(0.0, candidate.or_output_cost),
                                },
                            },
                        }
                        try:
                            async with client.stream(
                                "POST", f"{settings.openrouter_base_url}/chat/completions",
                                headers=headers, json=body,
                            ) as resp:
                                if resp.status_code >= 400:
                                    raw_error = (await resp.aread()).decode(errors="ignore")
                                    provider_error = f"OpenRouter {resp.status_code}: {raw_error[:200]}"
                                    candidate.error_count += 1
                                    candidate.last_provider_error = provider_error[:1000]
                                    continue
                                effective_model = candidate
                                succeeded = True
                                yield f"data: {json.dumps({'type': 'route', 'intent': 'text', 'requested_model': req.requested_model or req.model, 'effective_model': candidate.or_model_id, 'effective_model_name': candidate.name}, ensure_ascii=False)}\n\n"
                                async for line in resp.aiter_lines():
                                    if not line.startswith("data: "):
                                        continue
                                    data_str = line[6:].strip()
                                    if not data_str or data_str == "[DONE]":
                                        continue
                                    try:
                                        chunk = json.loads(data_str)
                                    except json.JSONDecodeError:
                                        continue
                                    if "error" in chunk:
                                        provider_error = str(chunk["error"])[:500]
                                        continue
                                    if "choices" in chunk and chunk["choices"]:
                                        delta = chunk["choices"][0].get("delta", {})
                                        token = delta.get("content", "")
                                        if token:
                                            full_content += token
                                            yield f"data: {json.dumps({'type': 'content', 'content': token})}\n\n"
                                    if "usage" in chunk:
                                        final_usage = chunk["usage"]
                                if full_content or final_usage:
                                    break
                                succeeded = False
                                candidate.error_count += 1
                                candidate.last_provider_error = "OpenRouter вернул пустой ответ"
                        except httpx.HTTPError as exc:
                            provider_error = str(exc)[:500]
                            candidate.error_count += 1
                            candidate.last_provider_error = provider_error
                            # Never concatenate two providers into one answer.
                            # If a stream already produced content, return that
                            # partial answer without charging instead of retrying.
                            if full_content:
                                succeeded = True
                                effective_model = candidate
                                final_usage = {}
                                break
            if not succeeded:
                await record_server_event(
                    db, user, "generation_failed", template_id=req.template_id,
                    task_type=req.task_type or "text", model=req.model,
                    metadata={"error_code": "provider_unavailable"},
                )
                await db.commit()
                friendly = "AI-сервис сейчас не отвечает. Мы не списали кредиты — повторите запрос через минуту."
                if not settings.openrouter_api_key:
                    friendly = provider_error
                yield f"data: {json.dumps({'type': 'error', 'content': friendly}, ensure_ascii=False)}\n\n"
                yield f"data: {json.dumps({'type': 'done', 'credits_spent': 0})}\n\n"
                yield "data: [DONE]\n\n"
                return

            # Deduct credits after streaming
            credits_spent = 0
            effective_info = _to_model_info(effective_model)
            if user and (effective_info.price_per_1k_input > 0 or effective_info.fixed_price > 0) and final_usage:
                if effective_info.fixed_price > 0:
                    credits_spent = max(1, math.ceil(effective_info.fixed_price))
                else:
                    input_tokens = final_usage.get("prompt_tokens", 0)
                    output_tokens = final_usage.get("completion_tokens", 0)
                    provider_cost = final_usage.get("cost")
                    if provider_cost is not None:
                        context = await pricing_context(db)
                        credits_spent = max(1, int(credits_for_provider_cost(float(provider_cost), context, whole=True)))
                    elif input_tokens or output_tokens:
                        credits_spent = max(1, math.ceil(
                            input_tokens * effective_info.price_per_1k_input / 1000
                            + output_tokens * effective_info.price_per_1k_output / 1000
                        ))
                    if credits_spent > 0 and not await _charge_credits(db, user.id, credits_spent, f"Чат: {effective_model.or_model_id}"):
                        logger.warning("Concurrent credit charge rejected for user=%s", user.id)
                        credits_spent = 0

            result_kind = "document" if "document" in (req.task_type or "") else "text"
            await record_server_event(
                db, user, "result_success", template_id=req.template_id,
                task_type=req.task_type or "text", model=effective_model.or_model_id,
                metadata={"result_kind": result_kind, "credits": credits_spent, "provider_cost_usd": final_usage.get("cost", "")},
            )
            await db.commit()

            yield f"data: {json.dumps({'type': 'done', 'credits_spent': credits_spent})}\n\n"
            yield "data: [DONE]\n\n"

        return StreamingResponse(event_stream(), media_type="text/event-stream")

    # ── Non-streaming path returns SSE too (for fixed_price models) ──
    async def non_streaming_as_sse():
        if not settings.openrouter_api_key:
            await record_server_event(
                db, user, "generation_failed", template_id=req.template_id,
                task_type=req.task_type or "text", model=req.model,
                metadata={"error_code": "provider_not_configured"},
            )
            await db.commit()
            yield f"data: {json.dumps({'type': 'error', 'content': 'AI-сервис временно не настроен. Попробуйте позже или обратитесь в поддержку.'}, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
            return
        try:
            async with httpx.AsyncClient(timeout=120.0, proxy=proxy) as client:
                response = await client.post(
                    f"{settings.openrouter_base_url}/chat/completions",
                    headers=headers,
                    json={
                        **base_body,
                        "provider": {
                            "sort": "price",
                            "max_price": {
                                "prompt": max(0.0, model.or_input_cost),
                                "completion": max(0.0, model.or_output_cost),
                            },
                        },
                    },
                )
        except httpx.HTTPError:
            await record_server_event(
                db, user, "generation_failed", template_id=req.template_id,
                task_type=req.task_type or "text", model=req.model,
                metadata={"error_code": "provider_network"},
            )
            await db.commit()
            yield f"data: {json.dumps({'type': 'error', 'content': 'AI-сервис сейчас не отвечает. Кредиты не списаны.'}, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
            return

        if response.status_code != 200:
            await record_server_event(
                db, user, "generation_failed", template_id=req.template_id,
                task_type=req.task_type or "text", model=req.model,
                metadata={"error_code": f"provider_{response.status_code}"},
            )
            await db.commit()
            yield f"data: {json.dumps({'type': 'error', 'content': 'AI-сервис отклонил запрос. Кредиты не списаны — попробуйте другую модель.'}, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
            return

        data = response.json()
        choice = data["choices"][0]
        msg = choice["message"]
        content = msg.get("content") or msg.get("text", "") or ""
        # Handle image generation models (Gemini, etc.) — images come in separate field
        images = msg.get("images")
        if images and isinstance(images, list) and len(images) > 0:
            img_url = images[0].get("image_url", {}).get("url", "") if isinstance(images[0], dict) else ""
            if img_url:
                if content:
                    content += f"\n\n![generated]({img_url})"
                else:
                    content = img_url
        if not content:
            content = f"[{model_info.name} не вернул ответ]"

        # Deduct credits
        credits_spent = 0
        if user and (model_info.price_per_1k_input > 0 or model_info.fixed_price > 0):
            if model_info.fixed_price > 0:
                credits_spent = max(1, math.ceil(model_info.fixed_price))
            else:
                input_tokens = data.get("usage", {}).get("prompt_tokens", 0)
                output_tokens = data.get("usage", {}).get("completion_tokens", 0)
                provider_cost = data.get("usage", {}).get("cost")
                if provider_cost is not None:
                    context = await pricing_context(db)
                    credits_spent = max(1, int(credits_for_provider_cost(float(provider_cost), context, whole=True)))
                else:
                    credits_spent = max(1, math.ceil(
                        input_tokens * model_info.price_per_1k_input / 1000
                        + output_tokens * model_info.price_per_1k_output / 1000
                    ))
            if credits_spent > 0:
                if not await _charge_credits(db, user.id, credits_spent, f"Чат: {req.model}"):
                    credits_spent = 0

        result_kind = "document" if "document" in (req.task_type or "") else "text"
        await record_server_event(
            db, user, "result_success", template_id=req.template_id,
            task_type=req.task_type or "text", model=req.model,
            metadata={"result_kind": result_kind, "credits": credits_spent, "provider_cost_usd": data.get("usage", {}).get("cost", "")},
        )
        await db.commit()

        yield f"data: {json.dumps({'type': 'content', 'content': content})}\n\n"
        yield f"data: {json.dumps({'type': 'done', 'credits_spent': credits_spent})}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(non_streaming_as_sse(), media_type="text/event-stream")


# ──────────────── Session sync (cross-device history) ────────────────


@router.get("/sessions")
async def get_sessions(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get all sessions for current user, newest first."""
    from app.models.chat_session import ChatSession as ChatSessionModel
    cutoff = datetime.now(timezone.utc) - timedelta(days=settings.history_retention_days)
    result = await db.execute(
        select(ChatSessionModel)
        .where(ChatSessionModel.user_id == user.id, ChatSessionModel.updated_at >= cutoff)
        .order_by(ChatSessionModel.updated_at.desc())
    )
    return [s.to_dict() for s in result.scalars().all()]


def _query_content(content) -> tuple[str, bool]:
    """Keep prompt text searchable without copying embedded data URLs."""
    if isinstance(content, str):
        return content.strip(), False
    if not isinstance(content, list):
        return str(content or "").strip(), False
    texts: list[str] = []
    has_attachments = False
    for part in content:
        if not isinstance(part, dict):
            continue
        if part.get("type") == "text":
            texts.append(str(part.get("text", "")))
        elif part.get("type") in {"image_url", "video_url", "file"}:
            has_attachments = True
    return "\n".join(texts).strip(), has_attachments


def _query_model(messages: list, index: int) -> str | None:
    message = messages[index] if index < len(messages) and isinstance(messages[index], dict) else {}
    for candidate in (message, messages[index + 1] if index + 1 < len(messages) and isinstance(messages[index + 1], dict) else {}):
        model = candidate.get("effective_model") or candidate.get("requested_model")
        if model:
            return str(model)[:200]
    return None


async def _sync_user_queries(db: AsyncSession, session_id: str, user_id: int, messages: list) -> None:
    existing = (await db.execute(
        select(UserQuery).where(UserQuery.session_id == session_id)
    )).scalars().all()
    existing_by_index = {item.message_index: item for item in existing}
    current_indexes: set[int] = set()

    for index, message in enumerate(messages):
        if not isinstance(message, dict) or message.get("role") != "user":
            continue
        current_indexes.add(index)
        content, has_attachments = _query_content(message.get("content", ""))
        item = existing_by_index.get(index)
        if item is None:
            item = UserQuery(session_id=session_id, user_id=user_id, message_index=index, content=content)
            db.add(item)
        item.content = content
        item.model = _query_model(messages, index)
        item.has_attachments = has_attachments

    stale_ids = [item.id for item in existing if item.message_index not in current_indexes]
    if stale_ids:
        await db.execute(delete(UserQuery).where(UserQuery.id.in_(stale_ids)))


@router.put("/sessions")
async def save_session(
    req: SessionSaveRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create or update a session on the server."""
    from app.models.chat_session import ChatSession as ChatSessionModel
    import json

    existing = await db.get(ChatSessionModel, req.id)
    if existing:
        if existing.user_id != user.id:
            raise HTTPException(status_code=403, detail="Session belongs to another user")
        existing.title = req.title[:200]
        existing.messages = json.dumps(req.messages, ensure_ascii=False)
    else:
        session = ChatSessionModel(
            id=req.id,
            user_id=user.id,
            title=req.title[:200],
            messages=json.dumps(req.messages, ensure_ascii=False),
        )
        db.add(session)
    await _sync_user_queries(db, req.id, user.id, req.messages)
    await db.commit()
    return {"ok": True}


@router.delete("/sessions/{session_id}")
async def delete_session(
    session_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a session."""
    from app.models.chat_session import ChatSession as ChatSessionModel
    session = await db.get(ChatSessionModel, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.user_id != user.id:
        raise HTTPException(status_code=403, detail="Session belongs to another user")
    await db.execute(delete(UserQuery).where(UserQuery.session_id == session_id))
    await db.execute(delete(MessageFeedback).where(MessageFeedback.session_id == session_id))
    await db.delete(session)
    await db.commit()
    return {"ok": True}


@router.post("/feedback")
async def save_message_feedback(
    req: MessageFeedbackRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create or replace a like/dislike for an assistant response."""
    if req.feedback_type not in {"like", "dislike", "regenerate"}:
        raise HTTPException(status_code=422, detail="Unsupported feedback type")

    from app.models.chat_session import ChatSession as ChatSessionModel
    session = await db.get(ChatSessionModel, req.session_id)
    if session is None or session.user_id != user.id:
        raise HTTPException(status_code=404, detail="Session not found")

    item = (await db.execute(select(MessageFeedback).where(
        MessageFeedback.session_id == req.session_id,
        MessageFeedback.message_index == req.message_index,
        MessageFeedback.user_id == user.id,
    ))).scalar_one_or_none()
    if item is None:
        item = MessageFeedback(
            session_id=req.session_id,
            message_index=req.message_index,
            user_id=user.id,
            feedback_type=req.feedback_type,
            model=req.model,
        )
        db.add(item)
    else:
        item.feedback_type = req.feedback_type
        item.model = req.model
    await record_server_event(
        db, user, "result_reused" if req.feedback_type == "regenerate" else "result_feedback",
        task_type="chat", model=req.model,
        metadata={"status": req.feedback_type, "result_kind": "text"},
    )
    await db.commit()
    return {"ok": True}


FACTCHECK_MODEL = "openai/gpt-4o-mini"

FACTCHECK_SYSTEM_PROMPT = """Ты — факт-чекер. Проверяешь факты в ответах ИИ-ассистента.

Пользователь задал вопрос, ИИ-ассистент дал ответ. Проверь факты в ответе.

Правила:
1. Найди фактические ошибки, неточности, устаревшие данные
2. Отметь подтверждённые факты
3. Укажи неуверенные утверждения, которые требуют проверки
4. Оцени общую достоверность ответа в процентах (confidence)

Ответь ТОЛЬКО JSON без пояснений:
{
  "errors": [
    {"claim": "утверждение с ошибкой", "status": "incorrect", "correction": "как правильно"}
  ],
  "confidence": 85,
  "verified_claims": [
    {"claim": "подтверждённый факт", "status": "correct", "correction": null}
  ],
  "details": "краткое резюме проверки на русском"
}"""


@router.post("/factcheck", response_model=FactCheckResponse)
async def factcheck(
    req: FactCheckRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Check facts in an AI response using another model. Cost: 1 credit."""
    # Credit check
    if user.credits < 1:
        raise HTTPException(status_code=402, detail="Недостаточно кредитов. Пополните баланс.")

    headers = {
        "Authorization": f"Bearer {settings.openrouter_api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://ai-sphere.ru",
        "X-Title": "AI-Sphere",
    }

    body = {
        "model": FACTCHECK_MODEL,
        "messages": [
            {"role": "system", "content": FACTCHECK_SYSTEM_PROMPT},
            {"role": "user", "content": f"Вопрос пользователя: {req.prompt}\n\nОтвет ИИ:\n{req.response}"},
        ],
        "max_tokens": 2048,
        "temperature": 0.1,
    }

    proxy = settings.openrouter_proxy or None
    try:
        async with httpx.AsyncClient(timeout=60.0, proxy=proxy) as client:
            response = await client.post(
                f"{settings.openrouter_base_url}/chat/completions",
                headers=headers,
                json=body,
            )

        if response.status_code != 200:
            return FactCheckResponse(
                errors=[],
                confidence=50,
                verified_claims=[],
                details=f"Ошибка проверки: {response.status_code}",
            )

        data = response.json()
        content = data["choices"][0]["message"].get("content", "{}")

        # Extract JSON from response (handles markdown-wrapped JSON)
        json_match = re.search(r"\{.*\}", content, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            fc_response = FactCheckResponse(
                errors=[FactCheckClaim(**e) for e in result.get("errors", [])],
                confidence=result.get("confidence", 50),
                verified_claims=[FactCheckClaim(**c) for c in result.get("verified_claims", [])],
                details=result.get("details", "Проверка выполнена."),
            )
        else:
            fc_response = FactCheckResponse(details="Не удалось распарсить результат проверки.")

        if not await _charge_credits(db, user.id, 1, f"Факт-чек: {req.model_id[:30]}"):
            raise HTTPException(status_code=402, detail="Недостаточно кредитов")

        return fc_response
    except HTTPException:
        raise
    except Exception as e:
        return FactCheckResponse(
            errors=[],
            confidence=50,
            verified_claims=[],
            details=f"Ошибка проверки: {str(e)[:200]}",
        )
