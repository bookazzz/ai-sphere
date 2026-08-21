#!/usr/bin/env python3
"""
Build clustered competitor keywords Excel for ai-sphere.ru
Sources: Keys.so query reports + competitor domain reports + industry knowledge
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# === RAW KEYWORD DATA FROM KEYS.SO ===

# Query: "нейросеть онлайн" — дополняющие фразы + органическая выдача + контекст
keywords_neuro_online = [
    # Related phrases from дополняющие фразы
    ("нейросеть онлайн", 409554, 11440),
    ("онлайн нейросеть", 406327, 11130),
    ("бесплатная нейросеть онлайн", 301209, 416),
    ("ии онлайн бесплатно", 231309, 7068),
    ("онлайн ии", 382060, 29000),
    ("нейросеть онлайн бесплатно на русском", 39676, 9842),
    ("нейросеть сайт онлайн", 792, 3),
    # Organic competitors keywords
    ("giga.chat нейросеть", 0, 0),
    ("гигачат", 0, 0),
    ("нейросеть для поиска информации", 0, 0),
    ("как делать фотографии с помощью нейросети", 0, 0),
    # Ads keywords
    ("алиса нейросеть", 0, 0),
    ("rugpt нейросеть", 0, 0),
    ("geekbot нейросеть", 0, 0),
    ("robotext нейросеть", 0, 0),
]

# Query: "чат gpt" — high-volume
keywords_chatgpt = [
    ("чат gpt", 2382010, 1797760),
    ("чат гпт", 0, 0),
    ("чат gpt т", 116, 31),
    ("чатгтп", 341, 0),
    ("gpt чат", 0, 0),
    ("chat gpt на русском", 0, 0),
    ("chatgpt онлайн", 0, 0),
    ("чат gpt бесплатно", 0, 0),
    ("chatgpt без впн", 0, 0),
    ("chatgpt россия", 0, 0),
]

# Query: "deepseek"
keywords_deepseek = [
    ("deepseek", 197148, 37515),
    ("deepseek чат", 3328, 940),
    ("deep seek chat", 1981, 1101),
    ("dipseek", 16727, 14711),
    ("http deepseek com", 261, 101),
    ("deepseek на русском", 0, 0),
    ("deepseek без впн", 0, 0),
    ("deepseek войти", 0, 0),
    ("deepseek нейросеть", 0, 0),
    ("deepseek desktop", 0, 0),
    ("deep seek скачать", 20, 0),
    ("deepseeker", 0, 0),
    ("дипсик", 3139258, 2540127),
    ("дипсик что это", 2463, 927),
    ("дипсик ии", 7451, 3388),
    ("искусственный интеллект дипсик", 0, 0),
    ("депсик нейросеть на русском", 0, 0),
]

# Query: "нейросеть чат"
keywords_neuro_chat = [
    ("нейросеть чат", 283353, 3212),
    ("чат нейросеть", 354757, 4148),
    ("ии нейросеть чат", 1540, 170),
    ("нейросеть для общения бесплатно", 0, 0),
    ("чат с нейросетью", 0, 0),
]

# Query: "чат ии"
keywords_chat_ii = [
    ("чат ии", 402555, 47022),
    ("ии общение онлайн", 734, 14),
    ("ии чаты", 415457, 2817),
    ("чат ии онлайн", 0, 0),
]

# Competitor-specific keywords (from domain reports)
competitor_keywords = {
    "giga.chat": [
        "гигачат", "giga chat", "гига чат", "giga.chat нейросеть", 
        "сбер нейросеть", "кандинский", "gigachat", "сбер нейросеть онлайн",
        "нейросеть сбера", "гигачат онлайн", "giga chat войти"
    ],
    "rugpt.io": [
        "rugpt", "ru gpt", "русский gpt", "rugpt нейросеть",
        "нейросеть rugpt", "rugpt io", "gpt на русском",
        "нейросеть для текста", "текст нейросеть", "проверка орфографии"
    ],
    "robotext.io": [
        "robotext", "роботекст", "robotext io", "нейросеть robotext",
        "роботекст нейросеть", "рерайт текста", "проверка текста"
    ],
    "geekbot.ru": [
        "geekbot", "гикбот", "geekbot нейросеть", "geek bot",
        "исправить текст с помощью ии", "fix text with ai"
    ],
    "chadgpt.ru": [
        "chadgpt", "чад gpt", "chad gpt", "chadgpt нейросеть",
        "чат gpt на русском chadgpt"
    ],
    "max.ru": [
        "max нейросеть", "max ru нейросеть", "нейросеть max"
    ],
    "redigram.ru": [
        "redigram", "редиграм", "redigram нейросеть", "ии помощник"
    ],
    "cropmedia.ru": [
        "cropmedia", "кропмедиа", "нейросеть cropmedia"
    ],
    "textplus.ru": [
        "textplus", "текст плюс", "рерайт текста онлайн"
    ],
    "retext.ai": [
        "retext", "ритект", "retext ai", "перефразирование текста"
    ]
}

# General AI chat / aggregator keywords
general_ai_keywords = [
    ("нейросеть онлайн чат", 0, 0),
    ("искусственный интеллект онлайн чат", 0, 0),
    ("ai чат на русском", 0, 0),
    ("чат бот нейросеть", 0, 0),
    ("нейросеть дипсик чат", 0, 0),
    ("чат с ии бесплатно", 0, 0),
    ("gpt чат бот", 0, 0),
    ("claude чат", 0, 0),
    ("gemini чат", 0, 0),
    ("gpt агрегатор", 0, 0),
    ("нейросеть агрегатор", 0, 0),
    ("чат нейросетей", 0, 0),
    ("агрегатор нейросетей", 0, 0),
    ("все нейросети в одном месте", 0, 0),
    ("получить доступ к нейросетям", 0, 0),
    ("нейросети без впн", 0, 0),
    ("нейросети россия", 0, 0),
    ("chatgpt без впн", 0, 0),
    ("deepseek на русском", 0, 0),
    ("chatgpt без ограничений", 0, 0),
    ("нейросети на русском языке", 0, 0),
    ("оплата нейросетей картой рф", 0, 0),
    ("нейросеть сбп", 0, 0),
    ("бот с нейросетями", 0, 0),
    ("telegram бот нейросеть", 0, 0),
    ("нейросеть в телеграм", 0, 0),
    ("чат гпт телеграм", 0, 0),
    ("chatgpt telegram bot", 0, 0),
    ("искусственный интеллект онлайн", 0, 0),
    ("ии ассистент", 0, 0),
    ("виртуальный помощник нейросеть", 0, 0),
]

# === CLUSTERING ===

clusters = {
    "Агрегаторы нейросетей (прямые конкуренты)": [
        "агрегатор нейросетей", "все нейросети в одном месте", "чат нейросетей",
        "gpt агрегатор", "нейросеть агрегатор", "получить доступ к нейросетям",
        "бот с нейросетями", "нейросети россия", "ai чат на русском",
        "giga.chat нейросеть", "гигачат", "gigachat", "сбер нейросеть",
        "chadgpt", "чад gpt", "chad gpt", "chadgpt нейросеть",
        "geekbot", "гикбот", "geekbot нейросеть",
        "max нейросеть", "redigram", "редиграм", "ии помощник",
        "cropmedia", "нейросеть cropmedia",
    ],
    "ChatGPT и GPT-модели": [
        "чат gpt", "чат гпт", "чат gpt т", "чатгтп", "gpt чат",
        "chat gpt на русском", "chatgpt онлайн", "чат gpt бесплатно",
        "chatgpt без впн", "chatgpt россия", "chatgpt telegram bot",
        "чат гпт телеграм", "chatgpt без ограничений",
        "rugpt", "ru gpt", "русский gpt", "rugpt нейросеть",
        "gpt на русском", "чат gpt на русском chadgpt",
        "chatgpt telegram bot", "нейросеть rugpt",
    ],
    "DeepSeek": [
        "deepseek", "deepseek чат", "deep seek chat", "dipseek",
        "http deepseek com", "deepseek на русском", "deepseek без впн",
        "deepseek войти", "deepseek нейросеть", "deepseek desktop",
        "deep seek скачать", "deepseeker", "дипсик", "дипсик что это",
        "дипсик ии", "искусственный интеллект дипсик",
        "депсик нейросеть на русском", "нейросеть дипсик чат",
    ],
    "Нейросеть онлайн (общие запросы)": [
        "нейросеть онлайн", "онлайн нейросеть", "бесплатная нейросеть онлайн",
        "нейросеть онлайн бесплатно на русском", "нейросеть сайт онлайн",
        "ии онлайн бесплатно", "онлайн ии", "нейронка онлайн без регистрации",
        "нейросеть онлайн чат", "искусственный интеллект онлайн",
    ],
    "Чат с нейросетью / AI чат": [
        "нейросеть чат", "чат нейросеть", "ии нейросеть чат",
        "нейросеть для общения бесплатно", "чат с нейросетью",
        "чат ии", "ии общение онлайн", "ии чаты", "чат ии онлайн",
        "чат бот нейросеть", "чат с ии бесплатно",
    ],
    "Нейросети без VPN / для России": [
        "нейросети без впн", "chatgpt без впн", "deepseek на русском",
        "нейросети на русском языке", "оплата нейросетей картой рф",
        "нейросеть сбп", "нейросети россия", "нейросеть без впн",
    ],
    "Telegram-боты с нейросетями": [
        "telegram бот нейросеть", "нейросеть в телеграм", "чат гпт телеграм",
        "бот с нейросетями", "chatgpt telegram bot", "бот нейросеть тг",
    ],
    "Текстовые нейросети / рерайт": [
        "robotext", "роботекст", "robotext io", "robotext нейросеть",
        "рерайт текста", "textplus", "текст плюс", "рерайт текста онлайн",
        "retext", "retext ai", "перефразирование текста",
        "нейросеть для текста", "текст нейросеть", "проверка орфографии",
        "проверка текста", "исправить текст с помощью ии", "fix text with ai",
    ],
    "ИИ-помощники и ассистенты": [
        "ии ассистент", "виртуальный помощник нейросеть", "искусственный интеллект онлайн",
        "ии онлайн бесплатно", "алиса нейросеть", "gigachat",
        "нейросеть giga.chat", "сбер нейросеть онлайн",
        "geekbot нейросеть", "ии помощник redigram",
    ],
}

# Build the full keyword set with cluster assignment
all_keywords = []

def add_keywords(keyword_list, source):
    for item in keyword_list:
        if isinstance(item, tuple):
            kw, freq, exact_freq = item
        else:
            kw, freq, exact_freq = item, 0, 0
        
        # Find cluster
        kw_lower = kw.lower().strip()
        assigned_cluster = "Прочее"
        for cluster_name, cluster_kws in clusters.items():
            if any(ck.lower() in kw_lower or kw_lower in ck.lower() for ck in cluster_kws):
                assigned_cluster = cluster_name
                break
        
        # Find which competitors rank for this
        competing_sites = []
        for site, site_kws in competitor_keywords.items():
            if any(ck.lower() in kw_lower or kw_lower in ck.lower() for ck in site_kws):
                competing_sites.append(site)
        
        all_keywords.append({
            'keyword': kw,
            'frequency': freq,
            'exact_frequency': exact_freq,
            'cluster': assigned_cluster,
            'source_query': source,
            'competitors': ', '.join(competing_sites) if competing_sites else '—',
        })

# Add all keywords
add_keywords(keywords_neuro_online, "нейросеть онлайн")
add_keywords(keywords_chatgpt, "чат gpt")
add_keywords(keywords_deepseek, "deepseek / дипсик")
add_keywords(keywords_neuro_chat, "нейросеть чат")
add_keywords(keywords_chat_ii, "чат ии")
add_keywords(general_ai_keywords, "общие")

# Deduplicate by keyword
seen = set()
deduped = []
for item in all_keywords:
    key = item['keyword'].lower().strip()
    if key not in seen:
        seen.add(key)
        deduped.append(item)

sorted_kw = sorted(deduped, key=lambda x: x['frequency'], reverse=True)

# === CREATE EXCEL ===
wb = Workbook()
ws = wb.active
ws.title = "Конкуренты ключевые слова"

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
cluster_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: All keywords
headers = ['Ключевое слово', 'Частотность', 'Точная частотность', 'Кластер', 'Источник', 'Конкуренты']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for row_idx, item in enumerate(sorted_kw, 2):
    ws.cell(row=row_idx, column=1, value=item['keyword']).border = thin_border
    ws.cell(row=row_idx, column=2, value=item['frequency']).border = thin_border
    ws.cell(row=row_idx, column=3, value=item['exact_frequency']).border = thin_border
    ws.cell(row=row_idx, column=4, value=item['cluster']).border = thin_border
    ws.cell(row=row_idx, column=4).fill = cluster_fill
    ws.cell(row=row_idx, column=5, value=item['source_query']).border = thin_border
    ws.cell(row=row_idx, column=6, value=item['competitors']).border = thin_border

# Column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 40

# Sheet 2: By cluster
ws2 = wb.create_sheet("По кластерам")
headers2 = ['Кластер', 'Ключевых слов', 'Ключевые слова']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Group by cluster
cluster_groups = defaultdict(list)
for item in sorted_kw:
    cluster_groups[item['cluster']].append(item['keyword'])

row = 2
for cluster_name, kws in sorted(cluster_groups.items()):
    ws2.cell(row=row, column=1, value=cluster_name).border = thin_border
    ws2.cell(row=row, column=1).fill = cluster_fill
    ws2.cell(row=row, column=1).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=len(kws)).border = thin_border
    ws2.cell(row=row, column=3, value=', '.join(kws[:15]) + ('...' if len(kws) > 15 else '')).border = thin_border
    row += 1

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 80

# Sheet 3: By competitor
ws3 = wb.create_sheet("По конкурентам")
headers3 = ['Конкурент', 'Ключевые слова']
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

row = 2
for site, kws in sorted(competitor_keywords.items()):
    ws3.cell(row=row, column=1, value=site).border = thin_border
    ws3.cell(row=row, column=1).fill = cluster_fill
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=2, value=', '.join(kws)).border = thin_border
    row += 1

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 80

# Sheet 4: Competitor comparison
ws4 = wb.create_sheet("Сравнение конкурентов")
headers4 = ['Параметр', 'giga.chat', 'rugpt.io', 'robotext.io', 'geekbot.ru', 'chadgpt.ru', 'ai-sphere.ru']
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

comparison_data = [
    ('Запросов топ-1', '1 093', '774', '675', '100', '37', '0'),
    ('Запросов топ-3', '2 876', '2 792', '3 218', '567', '124', '0'),
    ('Запросов топ-5', '5 143', '5 040', '5 604', '1 260', '204', '0'),
    ('Запросов топ-10', '12 078', '12 038', '10 947', '3 898', '528', '0'),
    ('Запросов топ-50', '43 671', '64 570', '32 273', '21 877', '5 840', '0'),
    ('Органический трафик', '21 958', '4 295', '7 478', '1 975', '393', '0'),
    ('Страниц в выдаче', '355', '422', '91', '612', '23', '0'),
    ('DR (рейтинг домена)', '39', '28', '35', '25', '21', '0'),
    ('Бюджет Директа', '37 млн ₽', '9,9 млн ₽', '4,3 млн ₽', '—', '7 млн ₽', '81 ₽'),
    ('Объявлений в контексте', '524', '1 166', '206', '—', '220', '2'),
    ('Запросов в контексте', '13 732', '8 044', '2 476', '—', '3 327', '2'),
    ('Упоминаний в ИИ-ответах', '5 556', '619', '769', '535', '52', '0'),
]

for row_idx, (param, *values) in enumerate(comparison_data, 2):
    ws4.cell(row=row_idx, column=1, value=param).border = thin_border
    ws4.cell(row=row_idx, column=1).font = Font(bold=True)
    for col_idx, val in enumerate(values, 2):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

ws4.column_dimensions['A'].width = 30
for col in range(2, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# Freeze panes
ws.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'
ws4.freeze_panes = 'B2'

# Auto-filter
ws.auto_filter.ref = f"A1:F{len(sorted_kw)+1}"

output_path = '/root/ai-sphere/competitors_keywords.xlsx'
wb.save(output_path)
print(f"Excel saved to {output_path}")
print(f"Total keywords (deduped): {len(sorted_kw)}")
print(f"Clusters: {len(cluster_groups)}")
print(f"\nKeywords per cluster:")
for c, kws in sorted(cluster_groups.items()):
    print(f"  {c}: {len(kws)} keywords")
